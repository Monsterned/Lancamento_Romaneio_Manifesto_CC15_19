import re
import pyautogui
import os
import pyperclip
import pandas as pd
from datetime import datetime
from datetime import datetime, timedelta
import sys
from openpyxl import load_workbook

caminho = os.getcwd() 

def click_image(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def botao_copy(image_path,image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                center_x = position2.left + position2.width // 2
                center_y = position2.top + position2.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def status_manifesto(image_path,image_path2,image_path3, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Status Baixado.")
                pyautogui.sleep(1)
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Status Cadastrado.")
                pyautogui.sleep(3)
                click_image('baixar_manifesto.png')
                pyautogui.sleep(3)
                tela_baixa('tela_baixar_data.png')
                campo_data_baixa('campo_data_baixa.png','campo_data_baixa2.png')
                pyautogui.sleep(1)
                for i in range(16):
                    pyautogui.press("backspace")
                pyautogui.write(str(um_minuto_atras_formatado))
                pyautogui.sleep(1)
                click_image('ok5.png')
                pyautogui.sleep(1)
                click_image('ok.png')
                #click_image('ok4.png')
                pyautogui.sleep(1)
                click_image('nao.png')
                pyautogui.sleep(2)
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position3 = pyautogui.locateOnScreen(image_path3, confidence=confidence)
            if position3:
                print("Status Emitido.")
                pyautogui.sleep(3)
                click_image('baixar_manifesto.png')
                pyautogui.sleep(3)
                tela_baixa('tela_baixar_data.png')
                campo_data_baixa('campo_data_baixa.png','campo_data_baixa2.png')
                pyautogui.sleep(1)
                for i in range(16):
                    pyautogui.press("backspace")
                pyautogui.write(str(um_minuto_atras_formatado))
                pyautogui.sleep(1)
                click_image('ok5.png')
                pyautogui.sleep(1)
                click_image('ok.png')
                #click_image('ok4.png')
                pyautogui.sleep(1)
                click_image('nao.png')
                pyautogui.sleep(2)
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(3)

def click_info(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.moveTo(center_x, center_y)  # Movendo o cursor para a posição da imagem               
                pyautogui.moveRel(50, 0)  # Movendo o cursor para cima
                pyautogui.click()  # Clicando no local da imagem
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def quantidade_registros():
    click_image('atualizar_pesquisa.png')
    pyautogui.sleep(5)   
    click_registros('buscar_registro.png')
    pyperclip.copy('')  # Limpa o conteúdo copiado
    pyautogui.sleep(2)
    pyautogui.click(button='right')
    pyautogui.sleep(1)

    click_image('copy.png')
    #botao_copy('copy2.png','copy_cinza.png')
    pyautogui.sleep(0.5)

    click_image('buscar_filtros.png')

    try:
        text = pyperclip.paste()  # Obtém o texto da área de transferência
        qtd_registros = int(text)  # Converte o texto para um número inteiro
        print("Quantidade de manifestos:", qtd_registros)

    except ValueError:
        print("O conteúdo copiado não é um número válido.")
        qtd_registros = None  # Define como None se a conversão falhar
        
    except Exception as e:
        print("Ocorreu um erro:", str(e))
        qtd_registros = None  # Define como None para qualquer outro erro

    return qtd_registros

def tela_baixa(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            clique_baixa('sim_cancelamento.png')
            clique_baixa('yes.png')           
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def click(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            center_x = position.left + position.width // 2
            center_y = position.top + position.height // 2
            pyautogui.click(center_x, center_y)
            print("Imagem foi encontrada na tela.")
    except Exception as e:
        print("Imagem não encontrada na tela. Aguardando...")
    pyautogui.sleep(1)

def clique_baixa(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            center_x = position.left + position.width // 2
            center_y = position.top + position.height // 2
            pyautogui.click(center_x, center_y)
            click('baixar_manifesto.png')
            print("Clicou novamente na baixa.")
    except Exception as e:
        print("Imagem não encontrada na tela. Aguardando...")
    pyautogui.sleep(1)

def campo_data_baixa(image_path,image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                center_x = position2.left + position2.width // 2
                center_y = position2.top + position2.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def clicar_encerrar(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            center_x = position.left + position.width // 2
            center_y = position.top + position.height // 2
            pyautogui.click(center_x, center_y)
            click('encerrar_sefaz.png')
            print("Imagem foi encontrada na tela.")
    except Exception as e:
        print("Imagem não encontrada na tela. Aguardando...")
    pyautogui.sleep(1)

def click_encerrar(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            clicar_encerrar('ok.png')
            clicar_encerrar('yes.png')
            pyautogui.sleep(1)
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def verificar_campo(image_path,image_path2,image_path3,image_path4,confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2) 
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    image_path4 = os.path.join(current_dir, caminho_imagem, image_path4)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Campo encontrado.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Campo encontrado.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position3 = pyautogui.locateOnScreen(image_path3, confidence=confidence)
            if position3:
                print("Campo encontrado.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position4 = pyautogui.locateOnScreen(image_path4, confidence=confidence)
            if position4:
                print("Campo encontrado.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def click_registros(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.moveTo(center_x, center_y)  # Movendo o cursor para a posição da imagem               
                pyautogui.moveRel(90, 0)  # Movendo o cursor para cima
                pyautogui.click(clicks=2)  # Clicando no local da imagem
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

mudou = False
Planilha_Manifesto = pd.read_excel("planilha_romaneio.xlsx")

click_image('faturamento.png')
pyautogui.sleep(0.5)
click_image('faturamento_movimentacao.png')
pyautogui.sleep(0.5)
click_image('faturamento_movimentacao_manifesto.png')
pyautogui.sleep(0.5)
click_image('faturamento_movimentacao_manifesto_emissao.png')
pyautogui.sleep(5)

click_image('buscar.png')
pyautogui.sleep(3)
linha_especifica = 1

for i, linha in enumerate(Planilha_Manifesto.index):
    romaneio = Planilha_Manifesto.loc[linha, "ROMANEIO"]
    placa = Planilha_Manifesto.loc[linha, "PLACA"]
    agora = datetime.now()
    um_minuto_atras = agora - timedelta(minutes=1)
    agora_formatado = agora.strftime("%d/%m/%Y %H:%M")
    um_minuto_atras_formatado = um_minuto_atras.strftime("%d/%m/%Y%H:%M")
    linha_especifica += 1

    if isinstance(romaneio, str):
        # A variável romaneio é uma string
        print("romaneio é uma string, indo para o proximo")
        continue
    if mudou == True:
        click_info('mdfe_encerrado.png')
        pyautogui.sleep(1)
        click_image('opcao_nao.png')
        pyautogui.sleep(1)
        click_info('situacao_manifesto.png')
        click_image('situacao_manifesto_cadastrado.png')
        pyautogui.sleep(1)
        click_image('buscar_mais_filtros.png')
        pyautogui.sleep(1)
        click_info('buscar_mais_filtros_autorizado.png')
        pyautogui.sleep(1)
        for i in range(2):
            pyautogui.press('up')
        click_image('buscar_mais_filtros_autorizado_sim.png')
        pyautogui.sleep(1)
        click_image('buscar_filtros.png')
        mudou = False

    click_info('man_placa.png')
    pyautogui.sleep(1)
    for i in range(10):
        pyautogui.press('backspace')
    for i in range(10):
        pyautogui.press('del')
    pyautogui.write(str(placa))
    pyautogui.sleep(1)

    #PESQUISA MAN EM ABERTO NO ESTADO DE CADASTRADO
    if linha == 0:
        click_info('buscar_data_emissao.png')
        pyautogui.sleep(1)
        for i in range(10):
            pyautogui.press('backspace')
        for i in range(10):
            pyautogui.press('del')
        pyautogui.sleep(1)
        click_info('mdfe_encerrado.png')
        pyautogui.sleep(1)
        click_image('opcao_nao.png')
        pyautogui.sleep(1)
        click_info('situacao_manifesto.png')
        click_image('situacao_manifesto_cadastrado.png')
        pyautogui.sleep(1)
        click_image('buscar_mais_filtros.png')
        pyautogui.sleep(1)
        click_info('buscar_mais_filtros_autorizado.png')
        pyautogui.sleep(1)
        click_image('buscar_mais_filtros_autorizado_sim.png')
        pyautogui.sleep(1)
        click_image('buscar_filtros.png')
    
    qtd_registros = quantidade_registros()
    pyautogui.sleep(1)

    #PESQUISA MAN EM ABERTO NO ESTADO DE EMITIDO
    if qtd_registros is None:
        mudou = True
        click_info('situacao_manifesto.png')
        click_image('situacao_manifesto_emitido.png')
        qtd_registros = quantidade_registros()
        pyautogui.sleep(1)

    if qtd_registros is None:
        click_info('situacao_manifesto.png')
        click_image('situacao_manifesto_baixado.png')
        qtd_registros = quantidade_registros()
        pyautogui.sleep(1)

    if qtd_registros is None:
        click_info('situacao_manifesto.png')
        click_image('situacao_manifesto_emitido.png')
        pyautogui.sleep(1)
        click_info('mdfe_encerrado.png')
        pyautogui.sleep(1)
        click_image('opcao_sim.png')
        qtd_registros = quantidade_registros()
        pyautogui.sleep(1)

    if qtd_registros is None:
        click_info('situacao_manifesto.png')
        click_image('situacao_manifesto_cadastrado.png')
        qtd_registros = quantidade_registros()
        pyautogui.sleep(1)

    #manifesto nao enviado pro sefaz(cadastrado/mdfe nao autorizado)
    if qtd_registros is None:

        click_info('mdfe_encerrado.png')
        pyautogui.sleep(1)
        click_image('opcao_nao.png')
        click_image('buscar_mais_filtros.png')
        pyautogui.sleep(1)
        click_info('buscar_mais_filtros_autorizado.png')
        pyautogui.sleep(1)
        click_image('buscar_mais_filtros_autorizado_nao.png')
        pyautogui.sleep(1)
        click_image('buscar_filtros.png')
        #VER QUANTIDADE PARA CADASTRADO E EMITIDO
        qtd_registros = quantidade_registros()
        pyautogui.sleep(1)  
        if qtd_registros is not None:
            if qtd_registros > 0:
                caminho_do_arquivo = 'planilha_romaneio.xlsx'
                nome_da_aba = 'Sheet1'
                wb = load_workbook(caminho_do_arquivo)
                ws = wb[nome_da_aba]
                coluna_ost = 'H'  
                if linha > ws.max_row:
                    ws[coluna_ost + str(linha_especifica)] = 'ULTIMO MANIFESTO EM ABERTO'
                else:
                    ws[coluna_ost + str(linha_especifica)] = 'ULTIMO MANIFESTO EM ABERTO'
                wb.save(caminho_do_arquivo)
                wb.close()
                print('Manifesto em aberto')
                print('Ir para o próximo')  
                continue
    
    if qtd_registros is None:
        click_info('situacao_manifesto.png')
        click_image('situacao_manifesto_emitido.png')
        qtd_registros = quantidade_registros()
        if qtd_registros is not None:
            if qtd_registros > 0:
                caminho_do_arquivo = 'planilha_romaneio.xlsx'
                nome_da_aba = 'Sheet1'
                wb = load_workbook(caminho_do_arquivo)
                ws = wb[nome_da_aba]
                coluna_ost = 'H'  
                if linha > ws.max_row:
                    ws[coluna_ost + str(linha_especifica)] = 'ULTIMO MANIFESTO EM ABERTO'
                else:
                    ws[coluna_ost + str(linha_especifica)] = 'ULTIMO MANIFESTO EM ABERTO'
                wb.save(caminho_do_arquivo)
                wb.close()
                print('Manifesto em aberto')
                print('Ir para o próximo')  
                continue
    
    if qtd_registros is None:
        click_info('situacao_manifesto.png')
        click_image('situacao_manifesto_inconsistente.png')
        qtd_registros = quantidade_registros()
        if qtd_registros is not None:
            if qtd_registros > 0:
                caminho_do_arquivo = 'planilha_romaneio.xlsx'
                nome_da_aba = 'Sheet1'
                wb = load_workbook(caminho_do_arquivo)
                ws = wb[nome_da_aba]
                coluna_ost = 'H'  
                if linha > ws.max_row:
                    ws[coluna_ost + str(linha_especifica)] = 'ULTIMO MANIFESTO EM ABERTO INCONSISTENTE'
                else:
                    ws[coluna_ost + str(linha_especifica)] = 'ULTIMO MANIFESTO EM ABERTO INCONSISTENTE'
                wb.save(caminho_do_arquivo)
                wb.close()
                print('Manifesto em aberto')
                print('Ir para o próximo')  
                continue

    # Verifica se qtd_registros é None
    if qtd_registros is None:
        print('Ir para o próximo')
        pyautogui.sleep(0.5)
        continue

    click_image('caixinha_baixado.png')
    pyautogui.moveRel(0, - 100)  # Movendo o cursor para cima
    pyautogui.click()  # Clicando no local da imagem
    pyautogui.sleep(1)
    for i in range(10):
        pyautogui.press('down')
    click_image('selecionar_pesquisa.png')

    #copiar numero do manifesto
    pyperclip.copy('')
    click_info('numero_manifesto.png')
    pyautogui.sleep(2)
    pyautogui.click(button='right')
    pyautogui.sleep(1)

    #click_image('copy.png')
    botao_copy('copy2.png','copy_cinza.png')

    pyautogui.sleep(0.5)
    try:
        # Pega o conteúdo da área de transferência
        text = pyperclip.paste()
        manifesto = int(text)
        print("Número do MANIFESTO:", manifesto)
                    
    except ValueError:
        print("O conteúdo copiado não é um número válido.")
        manifesto = 'Erro ao copiar código do manifesto'
    except Exception as e:
        print("Ocorreu um erro:", str(e))


    status_manifesto('status_baixado.png','status_cadastrado.png','status_emitido.png',)


    click_image('encerrar_sefaz.png')
    click_encerrar('campo_encerramento_numero_manifesto.png')
    pyautogui.press('enter')
    pyautogui.write(str(manifesto))#numero do manifesto
    click_image('botao_encerrar.png')
    pyautogui.sleep(4)
    verificar_campo('encerrado.png','manifesto_ja_baixado.png','encerrado_secretaria.png','rejeicao_indevido.png')
    click_image('fechar_tela_encerramento.png')
    pyautogui.sleep(2)
    click_image('buscar.png')
    pyautogui.sleep(3)

click_image('buscar_voltar.png')


