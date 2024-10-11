import re
import pyautogui
import os
import pyperclip
import pandas as pd
from datetime import datetime
from datetime import datetime, timedelta
import sys
from openpyxl import load_workbook
import math
from unidecode import unidecode
import tkinter as tk
from tkinter import messagebox

caminho = os.getcwd() 

def click_image(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def efetuar(image_path,image_path2,image_path3,image_path4, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    image_path4 = os.path.join(current_dir, caminho_imagem, image_path4)
    status = False  # Variável de controle para saber se a imagem 4 foi encontrada
    efetuado = False
    while True:
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Imagem foi encontrada na tela.")
                click_image('yes2.png')
                pyautogui.sleep(2)
                click_image('no.png')
                pyautogui.sleep(2)
                click_image('ok2.png')
                pyautogui.sleep(5)
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")

        try:
            position3 = pyautogui.locateOnScreen(image_path3, confidence=confidence)
            if position3:
                click_image('no.png')
                pyautogui.sleep(10)

                while True:
                    try:
                        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
                        if position:
                            print("Imagem foi encontrada na tela.")
                            click('ok2.png')
                            caminho_do_arquivo = 'planilha_romaneio.xlsx'
                            nome_da_aba = 'Sheet1'
                            wb = load_workbook(caminho_do_arquivo)
                            ws = wb[nome_da_aba]
                            coluna_ost = 'H'  
                            if linha > ws.max_row:
                                ws[coluna_ost + str(linha_especifica)] = 'SEM FICHA DE TRAFEGO'
                            else:
                                ws[coluna_ost + str(linha_especifica)] = 'SEM FICHA DE TRAFEGO'
                            wb.save(caminho_do_arquivo)
                            wb.close()
                            status = True  # Define como None se a conversão falhar
                            efetuado = True
                            break
                    except Exception as e:
                        print("Imagem foi encontrada na tela.")
                    
                    try:
                        position4 = pyautogui.locateOnScreen(image_path4, confidence=confidence)
                        if position4:
                            click_image('ok2.png')
                            pyautogui.sleep(5)
                            print("Imagem foi encontrada na tela.")
                            efetuado = True
                            break
                    except:
                        print('')
            if efetuado == True:
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

    return status

def status_sefaz(image_path,image_path2,image_path3, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Manifesto autorizado no sefaz.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Manifesto não autorizado no sefaz.")
                caminho_do_arquivo = 'planilha_romaneio.xlsx'
                nome_da_aba = 'Sheet1'
                wb = load_workbook(caminho_do_arquivo)
                ws = wb[nome_da_aba]
                coluna_ost = 'H'  
                if linha > ws.max_row:
                    ws[coluna_ost + str(linha_especifica)] = 'SEM CIOT, NAO AUTORIZADO NO SEFAZ'
                else:
                    ws[coluna_ost + str(linha_especifica)] = 'SEM CIOT, NAO AUTORIZADO NO SEFAZ'
                wb.save(caminho_do_arquivo)
                wb.close()
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position3 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position3:
                print("Manifesto não autorizado no sefaz.")
                caminho_do_arquivo = 'planilha_romaneio.xlsx'
                nome_da_aba = 'Sheet1'
                wb = load_workbook(caminho_do_arquivo)
                ws = wb[nome_da_aba]
                coluna_ost = 'H'  
                if linha > ws.max_row:
                    ws[coluna_ost + str(linha_especifica)] = 'NAO AUTORIZADO NO SEFAZ'
                else:
                    ws[coluna_ost + str(linha_especifica)] = 'NAO AUTORIZADO NO SEFAZ'
                wb.save(caminho_do_arquivo)
                wb.close()
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def click_inclusao_romaneio(image_path,image_path2,image_path3, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    tentativa = False
    while True:
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Aviso de romaneio ja incluido")
                click_image('ok2.png')
                pyautogui.sleep(1)
                click_image('guia_geral_manifesto.png')
                pyautogui.sleep(5)
                click_info('tipo_manifesto.png')
                click_image('tipo_reentrega.png')
                pyautogui.sleep(3)
                click_image('guia_documentos.png')
                pyautogui.sleep(1)
                click_image('verdinho.png')
                pyautogui.sleep(3)
                if tentativa == False:
                    click_inclusao_romaneio('verdinho_todo_cinza.png', 'aviso_ja_entregue.png')
                else:
                    print('Ja incluido em manifesto do mesmo tipo')
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(3)
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        #INSERIR CLICAR NO YES ANTES DE MUDAR O TIPO
        pyautogui.sleep(3)
        try:
            position3 = pyautogui.locateOnScreen(image_path3, confidence=confidence)
            if position3:
                center_x = position3.left + position3.width // 2
                center_y = position3.top + position3.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                pyautogui.sleep(5)
                while True:
                    try:
                        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
                        if position:
                            center_x = position.left + position.width // 2
                            center_y = position.top + position.height // 2
                            pyautogui.click(center_x, center_y)
                            print("Imagem foi encontrada na tela.")
                            break
                    except Exception as e:
                        print("Imagem não encontrada na tela. Aguardando...")
                    pyautogui.sleep(1)
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)
        
def click_info(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.moveTo(center_x, center_y)  # Movendo o cursor para a posição da imagem               
                pyautogui.moveRel(50, 0)  # Movendo o cursor para cima
                pyautogui.click()  # Clicando no local da imagem
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def novo_lançamento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Liberado para inclusão.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def confirmacao_salvamento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Manifesto ainda nao salvo. Aguardando...")
        except Exception as e:
            print("Manifesto salvo")
            break
        pyautogui.sleep(1)

def verificacao_motorista(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Motorista encontrado na tela")
                break
        except Exception as e:
            print("Motorista nao encontrado na tela") 
            click('ok_marcado.png')   
        pyautogui.sleep(1)

def click_pasta_amarela(image_path,image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                center_x = position2.left + position2.width // 2
                center_y = position2.top + position2.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def click(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            center_x = position.left + position.width // 2
            center_y = position.top + position.height // 2
            pyautogui.click(center_x, center_y)
            print("Imagem foi encontrada na tela.")
    except Exception as e:
        print("Imagem não encontrada na tela. Aguardando...")
    pyautogui.sleep(1)

def salvar_base_romaneio():
    caminho_do_arquivo = 'planilha_romaneio.xlsx'
    nome_da_aba = 'Sheet1'
    wb = load_workbook(caminho_do_arquivo)
    ws = wb[nome_da_aba]
    coluna_ost = 'I'  
    if linha > ws.max_row:
        ws[coluna_ost + str(linha_especifica)] = 'CIDADE NAO CONSTA NA BASE OU SEM CADASTRO'
    else:
        ws[coluna_ost + str(linha_especifica)] = 'CIDADE NAO CONSTA NA BASE OU SEM CADASTRO'
    wb.save(caminho_do_arquivo)
    wb.close()

# Função para exibir uma mensagem de sucesso
def show_success_message(msg):
    root = tk.Tk()
    root.withdraw()  # Oculta a janela principal
    messagebox.showinfo("Lancamento Manifesto", msg)  # Exibe o título e a mensagem
    root.destroy()  # Fecha a janela após exibir a mensagem

Planilha_Manifesto = pd.read_excel("planilha_romaneio.xlsx")
Planilha_cidades = pd.read_excel("Linhas_Codhorario.xlsx")

linha_especifica = 1
for i, linha in enumerate(Planilha_Manifesto.index):
    placa = Planilha_Manifesto.loc[linha, "PLACA"]
    romaneio = Planilha_Manifesto.loc[linha, "ROMANEIO"]
    baixado = Planilha_Manifesto.loc[linha, "STATUS"]
    manifesto = Planilha_Manifesto.loc[linha, "MANIFESTO"]
    linha_especifica += 1
    if pd.notna(baixado):
        print('Existe status de pendencia')
        continue
    if pd.notna(manifesto):
        print('Manifesto ja existe')
        continue
    if isinstance(romaneio, str):
        # A variável romaneio é uma string
        print("romaneio é uma string, indo para o proximo")
        continue
    
    romaneio = int(float(romaneio))
    cidade = Planilha_Manifesto.loc[linha, "CIDADE"].upper()
    
    km = Planilha_Manifesto.loc[linha, "KM"]
    km =str(km)
    if km.count(",") == 2:
        # Remover as vírgulas dos milhares e substituir a última por um ponto
        km = km.replace(",", "", 1).replace(",", ".")
    km = float(km)
    km = round(km)  # Arredonda para o inteiro mais próximo
    motorista = Planilha_Manifesto.loc[linha, "COD MOTORISTA"]
    motorista = int(float(motorista))
    liberado = Planilha_Manifesto.loc[linha, "STATUS"]
    liberado = str(liberado)

    if 'Ç' in cidade:
        cidade = cidade.replace("Ç", "C")
    elif('ç' in cidade):
        cidade = cidade.replace("ç", "c")
    cidade = unidecode(cidade)
    if liberado.lower() == 'nan':
        
        if romaneio == 'AVISO DE MOTORISTA OU VEICULO':
            continue
        if cidade in Planilha_cidades['Cidade Destino'].values:
            linha_man = Planilha_cidades.loc[Planilha_cidades[Planilha_cidades['Cidade Destino'] == cidade].index.values, 'LINHA'].values[0]
            cod_horario = Planilha_cidades.loc[Planilha_cidades[Planilha_cidades['Cidade Destino'] == cidade].index.values, 'COD HORARIO'].values[0]
            
            if pd.isna(linha_man) or pd.isna(cod_horario):
                print('Sem linha e código de horário cadastrado')
                salvar_base_romaneio()
                continue
        else:
            print('Cidade não consta na base')
            salvar_base_romaneio()
            continue

        cod_horario = int(cod_horario)
        print(f'placa: {placa} romaneio:{romaneio} cidade:{cidade} linha:{linha_man} cod horario:{cod_horario}')
            
        click_image('incluir.png')
        novo_lançamento('campo_numero_manifesto.png')
        pyautogui.sleep(5)
        click_info('tipo_manifesto.png')
        click_image('tipo_distribuicao.png')
        click_info('man_placa.png')
        pyautogui.sleep(1)
        pyautogui.write(str(placa))
        pyautogui.press('tab')
        pyautogui.sleep(4)

        verificacao_motorista('man_motorista.png') #ver depois como se comporta
        click('ok_marcado.png')
        click('ok.png')
        click('ok2.png')
        click('ok4.png')
        click_info('motorista_manifesto.png')
        pyautogui.sleep(2)
        for i in range(10):
            pyautogui.press('backspace')
        for i in range(10):
            pyautogui.press('del')
        pyautogui.write(str(motorista))
        pyautogui.sleep(2)
        pyautogui.press('tab')
        pyautogui.sleep(4)
        click('ok_marcado.png')
        click('ok.png')
        click('ok2.png')
        click('ok4.png')
        pyautogui.sleep(1)
        click_info('linha.png')
        pyautogui.write(str(linha_man))
        pyautogui.press('tab')
        pyautogui.sleep(2)
        click_info('man_filial_destino.png')
        pyautogui.write('1')
        pyautogui.press('tab')
        pyautogui.sleep(2)
        click_info('cod_horario.png')
        for i in range(6):
            pyautogui.press('del')
        pyautogui.write(str(cod_horario))
        pyautogui.press('tab')
        pyautogui.sleep(2)
        click_info('km_inicial.png')
        pyautogui.sleep(2)
        pyautogui.write('0')
        pyautogui.press('tab')
        pyautogui.sleep(2)
        pyautogui.write(str(km))
        pyautogui.press('tab')
        pyautogui.sleep(2)
        click_image('salvar.png')
        pyautogui.sleep(2)
        confirmacao_salvamento('campo_numero_manifesto.png')
        click_info('numero_manifesto.png')
        pyautogui.sleep(2)
        pyautogui.click(button='right')
        pyautogui.sleep(1)
        click_image('copy.png')
        pyautogui.sleep(0.5)
        try:
            # Pega o conteúdo da área de transferência
            text = pyperclip.paste()
            manifesto = int(text)
            print("Número do MANIFESTO:", manifesto)
                        
        except ValueError:
            print("O conteúdo copiado não é um número válido.")
            manifesto = 'Erro ao copiar código do manifesto'
        except Exception as e:
            print("Ocorreu um erro:", str(e))

        # Definir o nome do arquivo Excel
        file_name = 'planilha_romaneio.xlsx'

        # Verificar se o arquivo já existe
        try:
            df = pd.read_excel(file_name)
        except FileNotFoundError:
            print(f"O arquivo {file_name} não foi encontrado.")
            df = pd.DataFrame(columns=['ROMANEIO', 'PLACA', 'MANIFESTO'])

        # Atualizar o valor da coluna 'MANIFESTO' na linha encontrada
        df.at[linha, 'MANIFESTO'] = manifesto

        # Salvar a planilha atualizada
        df.to_excel(file_name, index=False)
        print(f"Manifesto {manifesto} atualizado na linha {linha} da planilha {file_name}.")

        pyautogui.sleep(3)
        click_image('guia_documentos.png')
        pyautogui.sleep(1)
        click_pasta_amarela('pasta_amarela_manifesto.png', 'pasta_amarela_manifesto_marcada.png')
        pyautogui.sleep(2)
        click_info('tipo_de_documento.png')
        pyautogui.sleep(2)
        click_image('tipo_de_documento_romaneio.png')
        pyautogui.sleep(2)
        click_info('filial.png')
        pyautogui.sleep(1)
        pyautogui.write('1')
        pyautogui.sleep(1)
        pyautogui.press('tab')
        pyautogui.sleep(2)
        pyautogui.write(str(romaneio))
        pyautogui.press('tab')
        pyautogui.sleep(5)
        click_image('verdinho.png')

        tentativa = click_inclusao_romaneio('verdinho_todo_cinza.png', 'manifesto_mesmo_tipo.png','man_doc_yes.png')

        if tentativa == True:
            caminho_do_arquivo = 'planilha_romaneio.xlsx'
            nome_da_aba = 'Sheet1'
            wb = load_workbook(caminho_do_arquivo)
            ws = wb[nome_da_aba]
            coluna_ost = 'H'  
            if linha > ws.max_row:
                ws[coluna_ost + str(linha_especifica)] = 'DOCUMENTO JA INCLUIDO EM MANIFESTO DO MESMO TIPO'
            else:
                ws[coluna_ost + str(linha_especifica)] = 'DOCUMENTO JA INCLUIDO EM MANIFESTO DO MESMO TIPO'
            wb.save(caminho_do_arquivo)
            wb.close()
            continue

        pyautogui.sleep(1)
        click_image('guia_geral_manifesto.png')
        pyautogui.sleep(3)
        click_image('botao_recalculo.png')
        pyautogui.sleep(2)
        click_image('ok3.png')         
        #click_image('ok4.png')
        pyautogui.sleep(10)
        click_image('botao_efetuar_manifesto.png')
        pyautogui.sleep(4)

        status = efetuar('ficha_nao_aberta.png','ctes_mesmo_dia.png','deseja_apagar_valores.png','manifesto_efetuado.png')
        if status is True:
            status = False
            continue
        
        pyautogui.sleep(5)
        click_image('enviar_sefaz.png')
        pyautogui.sleep(5)
        click_image('enviar_sefaz_enviar.png')
        pyautogui.sleep(2)
        pyautogui.press('left')
        click_image('yes2.png')
        pyautogui.sleep(2)

        #EMISSAO DE MANIFESTO POSSIVEIS ERROS
        #FICHA,CIOT, SEFAZ
        status_sefaz('manifesto_autorizado_sefaz.png','sem_ciot.png','erro_sefaz_1.png')
        pyautogui.sleep(5)
        click_image('sair_tela_sefaz.png')
        pyautogui.sleep(5)

pyautogui.sleep(5)

click_image('voltar.png')

# Simulando o final do código
if __name__ == "_main_":
    # Aqui você pode adicionar seu código que será executado
    # e ao final, chamará a função para exibir a mensagem
    show_success_message("Robô finalizado com sucesso")